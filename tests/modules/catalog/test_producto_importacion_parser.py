"""Tests unitarios del parser de importación Excel."""
from io import BytesIO

from openpyxl import Workbook

from app.modules.catalog.infrastructure.producto_importacion_parser import (
    PRESENTACION_HEADERS,
    SHEET_PRESENTACIONES,
    SHEET_PRODUCTOS,
    TEMPLATE_HEADERS,
    ProductoImportacionParser,
)


def test_parsear_productos_extrae_filas():
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PRODUCTOS
    ws.append(TEMPLATE_HEADERS)
    ws.append(["SKU-1", "Tornillo", "", 2, 50, "7800001", 0])
    buffer = BytesIO()
    wb.save(buffer)

    parser = ProductoImportacionParser()
    filas = parser.parsear_productos(buffer.getvalue())

    assert len(filas) == 1
    assert filas[0]["sku"] == "SKU-1"
    assert filas[0]["nombre"] == "Tornillo"
    assert filas[0]["unidad_medida_id"] == 2


def test_parsear_presentaciones_extrae_filas():
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PRESENTACIONES
    ws.append(PRESENTACION_HEADERS)
    ws.append(["SKU-1", "Caja", "7800002", 12, 1200, ""])
    buffer = BytesIO()
    wb.save(buffer)

    parser = ProductoImportacionParser()
    filas = parser.parsear_presentaciones(buffer.getvalue())

    assert len(filas) == 1
    assert filas[0]["sku"] == "SKU-1"
    assert filas[0]["nombre_presentacion"] == "Caja"
    assert filas[0]["cantidad_contenida"] == 12


def test_generar_plantilla_incluye_hojas_referencia():
    from openpyxl import load_workbook

    unidad = type("U", (), {"id": 1, "codigo": "UN", "nombre": "Unidad"})()
    tipo = type("T", (), {"id": 3, "nombre": "General"})()
    parser = ProductoImportacionParser()

    contenido = parser.generar_plantilla([unidad], [tipo])

    loaded = load_workbook(BytesIO(contenido), read_only=True)
    assert SHEET_PRODUCTOS in loaded.sheetnames
    assert SHEET_PRESENTACIONES in loaded.sheetnames
    loaded.close()
