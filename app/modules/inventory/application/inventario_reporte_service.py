"""Exportación de reportes de inventario (stock por ubicación, movimientos)."""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.modules.inventory.domain.ports import IInventarioMovimientosQuery, IInventarioStockQuery

FormatoExport = Literal["xlsx", "pdf"]
MAX_FILAS_EXPORT = 50_000

STOCK_HEADERS = [
    "SKU",
    "Producto",
    "Bodega",
    "Zona",
    "Tipo zona",
    "Cantidad",
    "Unidad",
]

MOV_HEADERS = [
    "Fecha",
    "Tipo",
    "SKU",
    "Producto",
    "Cantidad",
    "Origen",
    "Destino",
    "Usuario",
    "Documento",
    "Observaciones",
]


class InventarioReporteService:
    def __init__(
        self,
        listar_stock: IInventarioStockQuery,
        listar_movimientos: IInventarioMovimientosQuery,
    ):
        self._listar_stock = listar_stock
        self._listar_movimientos = listar_movimientos

    async def _obtener_stock_filas(self, empresa_id: int, **kwargs: Any) -> list[dict]:
        resultado = await self._listar_stock.handle(
            empresa_id,
            pagina=1,
            por_pagina=MAX_FILAS_EXPORT,
            **kwargs,
        )
        total = resultado["total"]
        if total > MAX_FILAS_EXPORT:
            raise ValueError(
                f"Demasiados registros ({total}). Aplique filtros (bodega, producto o zona) "
                f"o exporte como máximo {MAX_FILAS_EXPORT:,} filas."
            )
        return resultado["stock"]

    async def _obtener_movimientos_filas(self, empresa_id: int, **kwargs: Any) -> list[dict]:
        resultado = await self._listar_movimientos.handle(
            empresa_id,
            pagina=1,
            por_pagina=MAX_FILAS_EXPORT,
            **kwargs,
        )
        total = resultado["total"]
        if total > MAX_FILAS_EXPORT:
            raise ValueError(
                f"Demasiados movimientos ({total}). Aplique filtros (producto o tipo) "
                f"o exporte como máximo {MAX_FILAS_EXPORT:,} filas."
            )
        return resultado["movimientos"]

    @staticmethod
    def _stamp() -> str:
        return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")

    @staticmethod
    def _estilar_encabezado_excel(ws, col_count: int) -> None:
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

    def _stock_a_filas(self, items: list[dict]) -> list[list[Any]]:
        return [
            [
                row.get("producto_sku") or "",
                row.get("producto_nombre") or "",
                row.get("bodega_nombre") or "",
                row.get("zona_nombre") or "",
                row.get("tipo_zona_nombre") or "",
                row.get("cantidad"),
                row.get("unidad_medida_nombre") or "",
            ]
            for row in items
        ]

    def _mov_a_filas(self, items: list[dict]) -> list[list[Any]]:
        return [
            [
                (row.get("creado_at") or "")[:19],
                row.get("tipo") or "",
                row.get("producto_sku") or "",
                row.get("producto_nombre") or "",
                row.get("cantidad"),
                row.get("zona_origen_nombre") or "",
                row.get("zona_destino_nombre") or "",
                row.get("usuario_email") or "",
                " ".join(
                    p
                    for p in [row.get("documento_tipo"), row.get("documento_folio")]
                    if p
                ),
                row.get("observaciones") or "",
            ]
            for row in items
        ]

    def _generar_excel(self, titulo_hoja: str, headers: list[str], filas: list[list[Any]]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = titulo_hoja[:31]
        ws.append(headers)
        self._estilar_encabezado_excel(ws, len(headers))
        for fila in filas:
            ws.append(fila)
        for col in ws.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_len + 2, 48)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _generar_pdf(
        self,
        titulo: str,
        headers: list[str],
        filas: list[list[Any]],
        landscape_mode: bool = True,
    ) -> bytes:
        buf = BytesIO()
        pagesize = landscape(A4) if landscape_mode else A4
        doc = SimpleDocTemplate(buf, pagesize=pagesize, leftMargin=24, rightMargin=24, topMargin=36, bottomMargin=24)
        styles = getSampleStyleSheet()
        story = [
            Paragraph(titulo, styles["Heading2"]),
            Spacer(1, 8),
            Paragraph(
                f"Generado: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} · {len(filas)} registros",
                styles["Normal"],
            ),
            Spacer(1, 12),
        ]
        data = [headers] + [[str(c) if c is not None else "" for c in row] for row in filas]
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1565C0")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.25, rl_colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#F5F5F5")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(table)
        doc.build(story)
        return buf.getvalue()

    async def exportar_stock(
        self,
        empresa_id: int,
        formato: FormatoExport,
        **listado_kwargs: Any,
    ) -> tuple[bytes, str, str]:
        items = await self._obtener_stock_filas(empresa_id, **listado_kwargs)
        filas = self._stock_a_filas(items)
        stamp = self._stamp()
        if formato == "xlsx":
            content = self._generar_excel("Stock ubicacion", STOCK_HEADERS, filas)
            return content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"stock_ubicacion_{stamp}.xlsx"
        content = self._generar_pdf("Stock por ubicación", STOCK_HEADERS, filas)
        return content, "application/pdf", f"stock_ubicacion_{stamp}.pdf"

    async def exportar_movimientos(
        self,
        empresa_id: int,
        formato: FormatoExport,
        **listado_kwargs: Any,
    ) -> tuple[bytes, str, str]:
        items = await self._obtener_movimientos_filas(empresa_id, **listado_kwargs)
        filas = self._mov_a_filas(items)
        stamp = self._stamp()
        if formato == "xlsx":
            content = self._generar_excel("Movimientos", MOV_HEADERS, filas)
            return content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"movimientos_inventario_{stamp}.xlsx"
        content = self._generar_pdf("Historial de movimientos", MOV_HEADERS, filas)
        return content, "application/pdf", f"movimientos_inventario_{stamp}.pdf"
