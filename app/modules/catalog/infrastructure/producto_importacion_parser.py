"""Parsing y generación de plantilla Excel para importación de productos."""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

SHEET_PRODUCTOS = "Productos"
SHEET_PRESENTACIONES = "Presentaciones"
SHEET_UNIDADES = "Unidades_medida"
SHEET_TIPOS = "Tipos_producto"
TEMPLATE_HEADERS = [
    "sku",
    "nombre",
    "id_tipo_producto",
    "unidad_base",
    "precio_costo",
    "codigo_barras",
    "serializado",
]
PRESENTACION_HEADERS = [
    "sku",
    "nombre_presentacion",
    "codigo_barras",
    "cantidad_contenida",
    "precio_venta",
    "precio_costo",
]

HEADER_A_COLUMNAS: dict[str, str] = {
    "sku": "sku",
    "nombre": "nombre",
    "precio costo": "precio_costo",
    "unidad medida id": "unidad_medida_id",
    "unidad base": "unidad_medida_id",
    "id unidad base": "unidad_medida_id",
    "id tipo producto": "tipo_producto_id",
    "tipo producto id": "tipo_producto_id",
    "codigo barras": "codigo_barras",
    "codigo de barras": "codigo_barras",
    "barcode": "codigo_barras",
    "ean": "codigo_barras",
    "serializado": "serializado",
}

HEADER_PRESENTACION: dict[str, str] = {
    "sku": "sku",
    "nombre presentacion": "nombre_presentacion",
    "presentacion": "nombre_presentacion",
    "nombre": "nombre_presentacion",
    "codigo barras": "codigo_barras",
    "codigo de barras": "codigo_barras",
    "barcode": "codigo_barras",
    "ean": "codigo_barras",
    "cantidad contenida": "cantidad_contenida",
    "factor conversion": "cantidad_contenida",
    "factor": "cantidad_contenida",
    "precio venta": "precio_venta",
    "precio costo": "precio_costo",
}

COLUMNAS_OBLIGATORIAS = {"sku", "nombre", "unidad_medida_id"}
COLUMNAS_OBLIGATORIAS_PRESENTACION = {"sku", "nombre_presentacion", "codigo_barras", "cantidad_contenida"}


class ProductoImportacionParser:
    def generar_plantilla(self, unidades: list[Any], tipos: list[Any]) -> bytes:
        wb = Workbook()
        ws_prod = wb.active
        ws_prod.title = SHEET_PRODUCTOS
        ws_prod.append(TEMPLATE_HEADERS)
        self._estilar_encabezado(ws_prod, len(TEMPLATE_HEADERS))
        ws_prod.append(["SKU001", "Producto ejemplo", "", 1, "", "7801111111111", 0])

        ws_pres = wb.create_sheet(SHEET_PRESENTACIONES)
        ws_pres.append(PRESENTACION_HEADERS)
        self._estilar_encabezado(ws_pres, len(PRESENTACION_HEADERS))
        ws_pres.append(["SKU001", "Unidad", "7801111111111", 1, 990, ""])
        ws_pres.append(["SKU001", "Display 12", "7802222222222", 12, 10900, ""])

        ws_um = wb.create_sheet(SHEET_UNIDADES)
        ws_um.append(["unidad_base", "codigo", "nombre"])
        self._estilar_encabezado(ws_um, 3)
        for u in unidades:
            ws_um.append([u.id, u.codigo, u.nombre])

        ws_tp = wb.create_sheet(SHEET_TIPOS)
        ws_tp.append(["id_tipo_producto", "nombre"])
        self._estilar_encabezado(ws_tp, 2)
        for t in tipos:
            ws_tp.append([t.id, t.nombre])

        instrucciones = wb.create_sheet("Instrucciones")
        instrucciones.append(["Importación masiva de productos y códigos de barras"])
        instrucciones.append([])
        instrucciones.append(["Hoja Productos (opcional si solo agrega presentaciones):"])
        instrucciones.append(["  sku, nombre, unidad_base — obligatorios"])
        instrucciones.append(["  codigo_barras — opcional; crea presentación «Unidad» factor 1"])
        instrucciones.append(["  serializado — 0 o 1 (opcional)"])
        instrucciones.append([])
        instrucciones.append(["Hoja Presentaciones (opcional):"])
        instrucciones.append(["  sku — debe existir en Productos o ya en el sistema"])
        instrucciones.append(["  nombre_presentacion, codigo_barras, cantidad_contenida — obligatorios"])
        instrucciones.append(["  precio_venta, precio_costo — opcionales"])
        instrucciones.append([])
        instrucciones.append(["Un mismo SKU puede tener varias filas en Presentaciones (caja, display, etc.)."])

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _estilar_encabezado(self, ws, columnas: int) -> None:
        fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font = Font(bold=True, color="FFFFFF")
        for col in range(1, columnas + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill
            cell.font = font

    def parsear_productos(self, contenido: bytes) -> list[dict[str, Any]]:
        wb = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
        if SHEET_PRODUCTOS not in wb.sheetnames:
            wb.close()
            return []

        ws = wb[SHEET_PRODUCTOS]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return []

        indices = self._indices_columnas(rows[0], HEADER_A_COLUMNAS)
        faltantes = COLUMNAS_OBLIGATORIAS - set(indices.keys())
        if faltantes:
            wb.close()
            labels = ", ".join(sorted(faltantes))
            raise ValueError(
                f"Faltan columnas obligatorias en Productos: {labels}. "
                "Use la plantilla actual (unidad_base, sku, nombre)."
            )

        filas: list[dict[str, Any]] = []
        for row_num, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            def val(key: str):
                idx = indices.get(key)
                if idx is None:
                    return None
                return row[idx] if idx < len(row) else None

            sku = self._celda_texto(val("sku"))
            nombre = self._celda_texto(val("nombre"))
            if not sku and not nombre:
                continue

            filas.append(
                {
                    "fila": row_num,
                    "sku": sku,
                    "nombre": nombre,
                    "unidad_medida_id": val("unidad_medida_id"),
                    "tipo_producto_id": val("tipo_producto_id"),
                    "precio_costo": val("precio_costo"),
                    "codigo_barras": val("codigo_barras"),
                    "serializado": val("serializado"),
                }
            )

        wb.close()
        return filas

    def parsear_presentaciones(self, contenido: bytes) -> list[dict[str, Any]]:
        wb = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
        if SHEET_PRESENTACIONES not in wb.sheetnames:
            wb.close()
            return []

        ws = wb[SHEET_PRESENTACIONES]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []

        indices = self._indices_columnas(rows[0], HEADER_PRESENTACION)
        faltantes = COLUMNAS_OBLIGATORIAS_PRESENTACION - set(indices.keys())
        if faltantes:
            labels = ", ".join(sorted(faltantes))
            raise ValueError(
                f"Faltan columnas obligatorias en Presentaciones: {labels}. "
                "Use: sku, nombre_presentacion, codigo_barras, cantidad_contenida."
            )

        filas: list[dict[str, Any]] = []
        for row_num, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            def val(key: str):
                idx = indices.get(key)
                if idx is None:
                    return None
                return row[idx] if idx < len(row) else None

            sku = self._celda_texto(val("sku"))
            nombre_pres = self._celda_texto(val("nombre_presentacion"))
            codigo = self._celda_texto(val("codigo_barras"))
            if not sku and not codigo and not nombre_pres:
                continue

            filas.append(
                {
                    "fila": row_num,
                    "sku": sku,
                    "nombre_presentacion": nombre_pres,
                    "codigo_barras": codigo,
                    "cantidad_contenida": val("cantidad_contenida"),
                    "precio_venta": val("precio_venta"),
                    "precio_costo": val("precio_costo"),
                }
            )
        return filas

    def _indices_columnas(self, header_row, mapping: dict[str, str]) -> dict[str, int]:
        indices: dict[str, int] = {}
        for i, cell in enumerate(header_row):
            norm = self._normalizar_header(cell)
            columna = mapping.get(norm)
            if columna and columna not in indices:
                indices[columna] = i
        return indices

    @staticmethod
    def _normalizar_header(value) -> str:
        if value is None:
            return ""
        return str(value).strip().lower().replace("_", " ").replace("  ", " ").strip()

    @staticmethod
    def _celda_texto(value) -> str:
        if value is None:
            return ""
        return str(value).strip()
