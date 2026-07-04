"""Generación de plantilla e importación masiva de productos desde Excel."""
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.infrastructure.models.usuario import Producto, ProductoPresentacion
from app.modules.catalog.infrastructure.producto_crud import ProductoCRUDRepository
from app.modules.catalog.infrastructure.producto_presentacion_crud import (
    ProductoPresentacionCRUDRepository,
)
from app.modules.catalog.infrastructure.tipo_producto_crud import TipoProductoCRUDRepository
from app.modules.catalog.infrastructure.unidad_medida_crud import UnidadMedidaCRUDRepository

SHEET_PRODUCTOS = "Productos"
SHEET_PRESENTACIONES = "Presentaciones"
SHEET_UNIDADES = "Unidades_medida"
SHEET_TIPOS = "Tipos_producto"
TEMPLATE_HEADERS = [
    "sku",
    "nombre",
    "id_tipo_producto",
    "unidad_base",
    "precio_costo",
    "codigo_barras",
    "serializado",
]
PRESENTACION_HEADERS = [
    "sku",
    "nombre_presentacion",
    "codigo_barras",
    "cantidad_contenida",
    "precio_venta",
    "precio_costo",
]
MAX_FILAS = 2000
MAX_PRESENTACIONES = 5000

HEADER_A_COLUMNAS: dict[str, str] = {
    "sku": "sku",
    "nombre": "nombre",
    "precio costo": "precio_costo",
    "unidad medida id": "unidad_medida_id",
    "unidad base": "unidad_medida_id",
    "id unidad base": "unidad_medida_id",
    "id tipo producto": "tipo_producto_id",
    "tipo producto id": "tipo_producto_id",
    "codigo barras": "codigo_barras",
    "codigo de barras": "codigo_barras",
    "barcode": "codigo_barras",
    "ean": "codigo_barras",
    "serializado": "serializado",
}

HEADER_PRESENTACION: dict[str, str] = {
    "sku": "sku",
    "nombre presentacion": "nombre_presentacion",
    "presentacion": "nombre_presentacion",
    "nombre": "nombre_presentacion",
    "codigo barras": "codigo_barras",
    "codigo de barras": "codigo_barras",
    "barcode": "codigo_barras",
    "ean": "codigo_barras",
    "cantidad contenida": "cantidad_contenida",
    "factor conversion": "cantidad_contenida",
    "factor": "cantidad_contenida",
    "precio venta": "precio_venta",
    "precio costo": "precio_costo",
}

COLUMNAS_OBLIGATORIAS = {"sku", "nombre", "unidad_medida_id"}
COLUMNAS_OBLIGATORIAS_PRESENTACION = {"sku", "nombre_presentacion", "codigo_barras", "cantidad_contenida"}


class ProductoImportacionService:
    def __init__(self, session: AsyncSession):
        self.session = session
        self.producto_repo = ProductoCRUDRepository(session)
        self.presentacion_repo = ProductoPresentacionCRUDRepository(session)
        self.unidad_repo = UnidadMedidaCRUDRepository(session)
        self.tipo_producto_repo = TipoProductoCRUDRepository(session)

    async def generar_plantilla(self, empresa_id: int) -> bytes:
        unidades, _ = await self.unidad_repo.listar(
            empresa_id=empresa_id,
            pagina=1,
            por_pagina=500,
            es_super_admin=False,
        )
        tipos, _ = await self.tipo_producto_repo.listar(
            empresa_id=empresa_id,
            pagina=1,
            por_pagina=500,
            es_super_admin=False,
        )

        wb = Workbook()
        ws_prod = wb.active
        ws_prod.title = SHEET_PRODUCTOS
        ws_prod.append(TEMPLATE_HEADERS)
        self._estilar_encabezado(ws_prod, len(TEMPLATE_HEADERS))
        ws_prod.append(["SKU001", "Producto ejemplo", "", 1, "", "7801111111111", 0])

        ws_pres = wb.create_sheet(SHEET_PRESENTACIONES)
        ws_pres.append(PRESENTACION_HEADERS)
        self._estilar_encabezado(ws_pres, len(PRESENTACION_HEADERS))
        ws_pres.append(["SKU001", "Unidad", "7801111111111", 1, 990, ""])
        ws_pres.append(["SKU001", "Display 12", "7802222222222", 12, 10900, ""])

        ws_um = wb.create_sheet(SHEET_UNIDADES)
        ws_um.append(["unidad_base", "codigo", "nombre"])
        self._estilar_encabezado(ws_um, 3)
        for u in unidades:
            ws_um.append([u.id, u.codigo, u.nombre])

        ws_tp = wb.create_sheet(SHEET_TIPOS)
        ws_tp.append(["id_tipo_producto", "nombre"])
        self._estilar_encabezado(ws_tp, 2)
        for t in tipos:
            ws_tp.append([t.id, t.nombre])

        instrucciones = wb.create_sheet("Instrucciones")
        instrucciones.append(["Importación masiva de productos y códigos de barras"])
        instrucciones.append([])
        instrucciones.append(["Hoja Productos (opcional si solo agrega presentaciones):"])
        instrucciones.append(["  sku, nombre, unidad_base — obligatorios"])
        instrucciones.append(["  codigo_barras — opcional; crea presentación «Unidad» factor 1"])
        instrucciones.append(["  serializado — 0 o 1 (opcional)"])
        instrucciones.append([])
        instrucciones.append(["Hoja Presentaciones (opcional):"])
        instrucciones.append(["  sku — debe existir en Productos o ya en el sistema"])
        instrucciones.append(["  nombre_presentacion, codigo_barras, cantidad_contenida — obligatorios"])
        instrucciones.append(["  precio_venta, precio_costo — opcionales"])
        instrucciones.append([])
        instrucciones.append(["Un mismo SKU puede tener varias filas en Presentaciones (caja, display, etc.)."])

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _estilar_encabezado(self, ws, columnas: int) -> None:
        fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font = Font(bold=True, color="FFFFFF")
        for col in range(1, columnas + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill
            cell.font = font

    async def importar_desde_excel(self, contenido: bytes, empresa_id: int) -> dict[str, Any]:
        unidades, _ = await self.unidad_repo.listar(
            empresa_id=empresa_id,
            pagina=1,
            por_pagina=500,
            es_super_admin=False,
        )
        tipos, _ = await self.tipo_producto_repo.listar(
            empresa_id=empresa_id,
            pagina=1,
            por_pagina=500,
            es_super_admin=False,
        )
        unidades_validas = {int(u.id) for u in unidades}
        tipos_validos = {int(t.id) for t in tipos}

        skus_bd, nombres_bd = await self._cargar_existentes(empresa_id)
        barcodes_bd = await self._cargar_barcodes_existentes(empresa_id)

        filas = self._parsear_productos(contenido)
        filas_pres_sheet = self._parsear_presentaciones(contenido)
        if not filas and not filas_pres_sheet:
            raise ValueError(
                "No se encontraron filas en las hojas 'Productos' ni 'Presentaciones'"
            )
        if len(filas) > MAX_FILAS:
            raise ValueError(f"Máximo {MAX_FILAS} productos por archivo")
        if len(filas_pres_sheet) > MAX_PRESENTACIONES:
            raise ValueError(f"Máximo {MAX_PRESENTACIONES} presentaciones por archivo")

        skus_archivo: set[str] = set()
        nombres_archivo: set[str] = set()
        barcodes_archivo: set[str] = set()
        validos: list[dict[str, Any]] = []
        errores: list[dict[str, Any]] = []

        if filas:
            for fila in filas:
                numero = fila["fila"]
                sku = fila.get("sku", "")
                nombre = fila.get("nombre", "")
                unidad_raw = fila.get("unidad_medida_id")
                tipo_raw = fila.get("tipo_producto_id")
                precio_raw = fila.get("precio_costo")
                codigo_barras = self._celda_texto(fila.get("codigo_barras"))
                serializado_raw = fila.get("serializado")

                errores_fila: list[str] = []

                if not sku:
                    errores_fila.append("sku es obligatorio")
                elif len(sku) > 100:
                    errores_fila.append("sku no puede superar 100 caracteres")
                elif sku in skus_archivo:
                    errores_fila.append(f"sku '{sku}' duplicado en el archivo")
                elif sku in skus_bd:
                    errores_fila.append(f"sku '{sku}' ya existe en la empresa")

                if not nombre:
                    errores_fila.append("nombre es obligatorio")
                elif len(nombre) > 255:
                    errores_fila.append("nombre no puede superar 255 caracteres")
                elif nombre in nombres_archivo:
                    errores_fila.append(f"nombre '{nombre}' duplicado en el archivo")
                elif nombre in nombres_bd:
                    errores_fila.append(f"nombre '{nombre}' ya existe en la empresa")

                unidad_id: int | None = None
                if unidad_raw is None or str(unidad_raw).strip() == "":
                    errores_fila.append("unidad_base (unidad_medida_id) es obligatorio")
                else:
                    try:
                        unidad_id = int(float(unidad_raw))
                        if unidad_id not in unidades_validas:
                            errores_fila.append(
                                f"unidad_base {unidad_id} no existe o no pertenece a su empresa"
                            )
                    except (TypeError, ValueError):
                        errores_fila.append("unidad_base debe ser un número entero (ID de Unidades_medida)")

                tipo_producto_id: int | None = None
                if tipo_raw is not None and str(tipo_raw).strip() != "":
                    try:
                        tipo_producto_id = int(float(tipo_raw))
                        if tipo_producto_id not in tipos_validos:
                            errores_fila.append(
                                f"id_tipo_producto {tipo_producto_id} no existe o no pertenece a su empresa"
                            )
                    except (TypeError, ValueError):
                        errores_fila.append("id_tipo_producto debe ser un número entero")

                precio_costo: float | None = None
                if precio_raw is not None and str(precio_raw).strip() != "":
                    try:
                        precio_costo = float(Decimal(str(precio_raw)))
                        if precio_costo < 0:
                            errores_fila.append("precio_costo no puede ser negativo")
                    except (InvalidOperation, ValueError):
                        errores_fila.append("precio_costo debe ser un número válido")

                serializado = False
                if serializado_raw is not None and str(serializado_raw).strip() != "":
                    txt = str(serializado_raw).strip().lower()
                    if txt in ("1", "true", "si", "sí", "yes"):
                        serializado = True
                    elif txt in ("0", "false", "no"):
                        serializado = False
                    else:
                        errores_fila.append("serializado debe ser 0 o 1")

                if codigo_barras:
                    if len(codigo_barras) > 100:
                        errores_fila.append("codigo_barras no puede superar 100 caracteres")
                    elif codigo_barras in barcodes_archivo:
                        errores_fila.append(f"codigo_barras '{codigo_barras}' duplicado en el archivo")
                    elif codigo_barras in barcodes_bd:
                        errores_fila.append(f"codigo_barras '{codigo_barras}' ya existe en la empresa")

                if errores_fila:
                    errores.append({"fila": numero, "sku": sku or None, "errores": errores_fila})
                    continue

                if codigo_barras:
                    barcodes_archivo.add(codigo_barras)

                skus_archivo.add(sku)
                nombres_archivo.add(nombre)
                validos.append(
                    {
                        "fila": numero,
                        "empresa_id": empresa_id,
                        "sku": sku,
                        "nombre": nombre,
                        "unidad_medida_id": unidad_id,
                        "tipo_producto_id": tipo_producto_id,
                        "precio_costo": precio_costo,
                        "serializado": serializado,
                        "codigo_barras": codigo_barras or None,
                        "activo": True,
                    }
                )

        creados = 0
        if validos:
            creados = await self.producto_repo.crear_masivo(
                [
                    {
                        "empresa_id": v["empresa_id"],
                        "sku": v["sku"],
                        "nombre": v["nombre"],
                        "unidad_medida_id": v["unidad_medida_id"],
                        "tipo_producto_id": v["tipo_producto_id"],
                        "precio_costo": v["precio_costo"],
                        "serializado": v["serializado"],
                        "activo": v["activo"],
                    }
                    for v in validos
                ]
            )

        skus_pres = {self._celda_texto(f.get("sku")) for f in filas_pres_sheet if self._celda_texto(f.get("sku"))}
        mapa_skus = await self._mapa_skus_empresa(empresa_id, skus_archivo | skus_bd | skus_pres)

        pres_auto: list[dict[str, Any]] = []
        for v in validos:
            if v.get("codigo_barras"):
                pres_auto.append(
                    {
                        "fila": v["fila"],
                        "sku": v["sku"],
                        "nombre_presentacion": "Unidad",
                        "codigo_barras": v["codigo_barras"],
                        "cantidad_contenida": Decimal("1"),
                        "precio_venta": None,
                        "precio_costo": v.get("precio_costo"),
                    }
                )

        filas_pres = filas_pres_sheet
        pres_validas, errores_pres = await self._validar_presentaciones(
            filas_pres + pres_auto,
            mapa_skus,
            barcodes_bd,
            empresa_id,
        )

        presentaciones_creadas = 0
        if pres_validas:
            presentaciones_creadas = await self._crear_presentaciones_masivo(pres_validas)

        return {
            "total_filas": len(filas) + len(filas_pres_sheet),
            "creados": creados,
            "presentaciones_creadas": presentaciones_creadas,
            "con_errores": len(errores) + len(errores_pres),
            "errores": errores,
            "errores_presentaciones": errores_pres,
        }

    async def _validar_presentaciones(
        self,
        filas: list[dict[str, Any]],
        mapa_skus: dict[str, int],
        barcodes_ocupados: set[str],
        empresa_id: int,
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        validas: list[dict[str, Any]] = []
        errores: list[dict[str, Any]] = []
        barcodes_en_archivo: set[str] = set()

        for fila in filas:
            numero = fila["fila"]
            sku = self._celda_texto(fila.get("sku"))
            nombre_pres = self._celda_texto(fila.get("nombre_presentacion"))
            codigo = self._celda_texto(fila.get("codigo_barras"))
            cant_raw = fila.get("cantidad_contenida")
            pv_raw = fila.get("precio_venta")
            pc_raw = fila.get("precio_costo")

            errores_fila: list[str] = []

            if not sku:
                errores_fila.append("sku es obligatorio")
            elif sku not in mapa_skus:
                errores_fila.append(f"sku '{sku}' no existe (créelo en Productos o importe antes)")

            if not nombre_pres:
                errores_fila.append("nombre_presentacion es obligatorio")
            if not codigo:
                errores_fila.append("codigo_barras es obligatorio")
            elif codigo in barcodes_en_archivo:
                errores_fila.append(f"codigo_barras '{codigo}' duplicado en el archivo")
            elif codigo in barcodes_ocupados:
                errores_fila.append(f"codigo_barras '{codigo}' ya existe en la empresa")

            cantidad: Decimal | None = None
            if cant_raw is None or str(cant_raw).strip() == "":
                errores_fila.append("cantidad_contenida es obligatoria")
            else:
                try:
                    cantidad = Decimal(str(cant_raw))
                    if cantidad <= 0:
                        errores_fila.append("cantidad_contenida debe ser mayor a cero")
                except (InvalidOperation, ValueError):
                    errores_fila.append("cantidad_contenida debe ser un número válido")

            precio_venta: float | None = None
            if pv_raw is not None and str(pv_raw).strip() != "":
                try:
                    precio_venta = float(Decimal(str(pv_raw)))
                    if precio_venta < 0:
                        errores_fila.append("precio_venta no puede ser negativo")
                except (InvalidOperation, ValueError):
                    errores_fila.append("precio_venta debe ser un número válido")

            precio_costo: float | None = None
            if pc_raw is not None and str(pc_raw).strip() != "":
                try:
                    precio_costo = float(Decimal(str(pc_raw)))
                    if precio_costo < 0:
                        errores_fila.append("precio_costo no puede ser negativo")
                except (InvalidOperation, ValueError):
                    errores_fila.append("precio_costo debe ser un número válido")

            if errores_fila:
                errores.append({"fila": numero, "sku": sku or None, "errores": errores_fila})
                continue

            barcodes_en_archivo.add(codigo)
            producto_id = mapa_skus[sku]
            producto = await self.producto_repo.obtener_por_id(producto_id, empresa_id)
            if not producto:
                errores.append({"fila": numero, "sku": sku, "errores": ["Producto no encontrado"]})
                continue

            if await self.presentacion_repo.obtener_por_nombre(producto_id, nombre_pres, solo_activas=True):
                errores.append(
                    {
                        "fila": numero,
                        "sku": sku,
                        "errores": [f"Ya existe la presentación '{nombre_pres}' para este producto"],
                    }
                )
                continue

            validas.append(
                {
                    "producto_id": producto_id,
                    "unidad_medida_id": producto.unidad_medida_id,
                    "nombre": nombre_pres,
                    "codigo_barras": codigo,
                    "cantidad_contenida": cantidad,
                    "precio_venta": precio_venta,
                    "precio_costo": precio_costo,
                }
            )

        return validas, errores

    async def _crear_presentaciones_masivo(self, items: list[dict[str, Any]]) -> int:
        if not items:
            return 0
        try:
            objetos = [
                ProductoPresentacion(
                    producto_id=item["producto_id"],
                    nombre=item["nombre"],
                    codigo_barras=item["codigo_barras"],
                    cantidad_contenida=item["cantidad_contenida"],
                    unidad_medida_id=item["unidad_medida_id"],
                    precio_costo=item.get("precio_costo"),
                    precio_venta=item.get("precio_venta"),
                    permite_venta_unidad=True,
                    permite_venta_presentacion=True,
                    activo=True,
                )
                for item in items
            ]
            self.session.add_all(objetos)
            await self.session.commit()
            return len(objetos)
        except Exception:
            await self.session.rollback()
            raise

    async def _mapa_skus_empresa(self, empresa_id: int, skus: set[str]) -> dict[str, int]:
        if not skus:
            return {}
        stmt = select(Producto.sku, Producto.id).where(
            Producto.empresa_id == empresa_id,
            Producto.sku.in_(list(skus)),
        )
        result = await self.session.execute(stmt)
        return {row[0]: int(row[1]) for row in result.all()}

    async def _cargar_barcodes_existentes(self, empresa_id: int) -> set[str]:
        stmt = (
            select(ProductoPresentacion.codigo_barras)
            .join(Producto, ProductoPresentacion.producto_id == Producto.id)
            .where(
                Producto.empresa_id == empresa_id,
                ProductoPresentacion.activo == True,
                ProductoPresentacion.codigo_barras.isnot(None),
            )
        )
        result = await self.session.execute(stmt)
        return {r[0] for r in result.all() if r[0]}

    async def _cargar_existentes(self, empresa_id: int) -> tuple[set[str], set[str]]:
        stmt = select(Producto.sku, Producto.nombre).where(Producto.empresa_id == empresa_id)
        result = await self.session.execute(stmt)
        rows = result.all()
        return {r[0] for r in rows if r[0]}, {r[1] for r in rows if r[1]}

    def _parsear_productos(self, contenido: bytes) -> list[dict[str, Any]]:
        wb = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
        if SHEET_PRODUCTOS not in wb.sheetnames:
            wb.close()
            return []

        ws = wb[SHEET_PRODUCTOS]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return []

        indices = self._indices_columnas(rows[0], HEADER_A_COLUMNAS)
        faltantes = COLUMNAS_OBLIGATORIAS - set(indices.keys())
        if faltantes:
            wb.close()
            labels = ", ".join(sorted(faltantes))
            raise ValueError(
                f"Faltan columnas obligatorias en Productos: {labels}. "
                "Use la plantilla actual (unidad_base, sku, nombre)."
            )

        filas: list[dict[str, Any]] = []
        for row_num, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            def val(key: str):
                idx = indices.get(key)
                if idx is None:
                    return None
                return row[idx] if idx < len(row) else None

            sku = self._celda_texto(val("sku"))
            nombre = self._celda_texto(val("nombre"))
            if not sku and not nombre:
                continue

            filas.append(
                {
                    "fila": row_num,
                    "sku": sku,
                    "nombre": nombre,
                    "unidad_medida_id": val("unidad_medida_id"),
                    "tipo_producto_id": val("tipo_producto_id"),
                    "precio_costo": val("precio_costo"),
                    "codigo_barras": val("codigo_barras"),
                    "serializado": val("serializado"),
                }
            )

        wb.close()
        return filas

    def _parsear_presentaciones(self, contenido: bytes) -> list[dict[str, Any]]:
        wb = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
        if SHEET_PRESENTACIONES not in wb.sheetnames:
            wb.close()
            return []

        ws = wb[SHEET_PRESENTACIONES]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []

        indices = self._indices_columnas(rows[0], HEADER_PRESENTACION)
        faltantes = COLUMNAS_OBLIGATORIAS_PRESENTACION - set(indices.keys())
        if faltantes:
            labels = ", ".join(sorted(faltantes))
            raise ValueError(
                f"Faltan columnas obligatorias en Presentaciones: {labels}. "
                "Use: sku, nombre_presentacion, codigo_barras, cantidad_contenida."
            )

        filas: list[dict[str, Any]] = []
        for row_num, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            def val(key: str):
                idx = indices.get(key)
                if idx is None:
                    return None
                return row[idx] if idx < len(row) else None

            sku = self._celda_texto(val("sku"))
            nombre_pres = self._celda_texto(val("nombre_presentacion"))
            codigo = self._celda_texto(val("codigo_barras"))
            if not sku and not codigo and not nombre_pres:
                continue

            filas.append(
                {
                    "fila": row_num,
                    "sku": sku,
                    "nombre_presentacion": nombre_pres,
                    "codigo_barras": codigo,
                    "cantidad_contenida": val("cantidad_contenida"),
                    "precio_venta": val("precio_venta"),
                    "precio_costo": val("precio_costo"),
                }
            )
        return filas

    def _indices_columnas(self, header_row, mapping: dict[str, str]) -> dict[str, int]:
        indices: dict[str, int] = {}
        for i, cell in enumerate(header_row):
            norm = self._normalizar_header(cell)
            columna = mapping.get(norm)
            if columna and columna not in indices:
                indices[columna] = i
        return indices

    @staticmethod
    def _normalizar_header(value) -> str:
        if value is None:
            return ""
        return str(value).strip().lower().replace("_", " ").replace("  ", " ").strip()

    @staticmethod
    def _celda_texto(value) -> str:
        if value is None:
            return ""
        return str(value).strip()
