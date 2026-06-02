"""Generación de plantilla e importación masiva de productos desde Excel."""
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.infrastructure.models.usuario import Producto
from app.infrastructure.repositories.producto_crud_repository import ProductoCRUDRepository
from app.infrastructure.repositories.tipo_producto_crud_repository import TipoProductoCRUDRepository
from app.infrastructure.repositories.unidadMedida_crud_repository import UnidadMedidaCRUDRepository

SHEET_PRODUCTOS = "Productos"
SHEET_UNIDADES = "Unidades_medida"
SHEET_TIPOS = "Tipos_producto"
TEMPLATE_HEADERS = ["sku", "nombre", "id_tipo_producto", "unidad_base", "precio_costo"]
MAX_FILAS = 2000

# Encabezados aceptados en fila 1 (compatibilidad plantilla antigua y alias en español).
HEADER_A_COLUMNAS: dict[str, str] = {
    "sku": "sku",
    "nombre": "nombre",
    "precio costo": "precio_costo",
    "unidad medida id": "unidad_medida_id",
    "unidad base": "unidad_medida_id",
    "id unidad base": "unidad_medida_id",
    "id tipo producto": "tipo_producto_id",
    "tipo producto id": "tipo_producto_id",
}

COLUMNAS_OBLIGATORIAS = {"sku", "nombre", "unidad_medida_id"}


class ProductoImportacionService:
    def __init__(self, session: AsyncSession):
        self.session = session
        self.producto_repo = ProductoCRUDRepository(session)
        self.unidad_repo = UnidadMedidaCRUDRepository(session)
        self.tipo_producto_repo = TipoProductoCRUDRepository(session)

    async def generar_plantilla(self, empresa_id: int) -> bytes:
        unidades, _ = await self.unidad_repo.listar(
            empresa_id=empresa_id,
            pagina=1,
            por_pagina=500,
            es_super_admin=False,
        )
        tipos, _ = await self.tipo_producto_repo.listar(
            empresa_id=empresa_id,
            pagina=1,
            por_pagina=500,
            es_super_admin=False,
        )

        wb = Workbook()
        ws_prod = wb.active
        ws_prod.title = SHEET_PRODUCTOS
        ws_prod.append(TEMPLATE_HEADERS)
        self._estilar_encabezado(ws_prod, len(TEMPLATE_HEADERS))
        ws_prod.append(["", "", "", "", ""])

        ws_um = wb.create_sheet(SHEET_UNIDADES)
        ws_um.append(["unidad_base", "codigo", "nombre"])
        self._estilar_encabezado(ws_um, 3)
        for u in unidades:
            ws_um.append([u.id, u.codigo, u.nombre])

        ws_tp = wb.create_sheet(SHEET_TIPOS)
        ws_tp.append(["id_tipo_producto", "nombre"])
        self._estilar_encabezado(ws_tp, 2)
        for t in tipos:
            ws_tp.append([t.id, t.nombre])

        instrucciones = wb.create_sheet("Instrucciones")
        instrucciones.append(["Importación masiva de productos"])
        instrucciones.append([])
        instrucciones.append(["Columnas hoja Productos:"])
        instrucciones.append(["  sku — código único del producto (obligatorio)"])
        instrucciones.append(["  nombre — descripción (obligatorio)"])
        instrucciones.append(["  id_tipo_producto — ID de Tipos_producto (opcional)"])
        instrucciones.append(["  unidad_base — ID de Unidades_medida, unidad de inventario (obligatorio)"])
        instrucciones.append(["  precio_costo — numérico (opcional)"])
        instrucciones.append([])
        instrucciones.append(["1. Use los IDs de las hojas Unidades_medida y Tipos_producto de su empresa."])
        instrucciones.append(["2. empresa_id y activo se asignan al importar según su sesión."])
        instrucciones.append(["3. No modifique los encabezados de la fila 1."])

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _estilar_encabezado(self, ws, columnas: int) -> None:
        fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font = Font(bold=True, color="FFFFFF")
        for col in range(1, columnas + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill
            cell.font = font

    async def importar_desde_excel(self, contenido: bytes, empresa_id: int) -> dict[str, Any]:
        unidades, _ = await self.unidad_repo.listar(
            empresa_id=empresa_id,
            pagina=1,
            por_pagina=500,
            es_super_admin=False,
        )
        tipos, _ = await self.tipo_producto_repo.listar(
            empresa_id=empresa_id,
            pagina=1,
            por_pagina=500,
            es_super_admin=False,
        )
        unidades_validas = {int(u.id) for u in unidades}
        tipos_validos = {int(t.id) for t in tipos}

        skus_bd, nombres_bd = await self._cargar_existentes(empresa_id)

        filas = self._parsear_productos(contenido)
        if not filas:
            raise ValueError("No se encontraron filas de productos en la hoja 'Productos'")

        if len(filas) > MAX_FILAS:
            raise ValueError(f"Máximo {MAX_FILAS} productos por archivo")

        skus_archivo: set[str] = set()
        nombres_archivo: set[str] = set()
        validos: list[dict[str, Any]] = []
        errores: list[dict[str, Any]] = []

        for fila in filas:
            numero = fila["fila"]
            sku = fila.get("sku", "")
            nombre = fila.get("nombre", "")
            unidad_raw = fila.get("unidad_medida_id")
            tipo_raw = fila.get("tipo_producto_id")
            precio_raw = fila.get("precio_costo")

            errores_fila: list[str] = []

            if not sku:
                errores_fila.append("sku es obligatorio")
            elif len(sku) > 100:
                errores_fila.append("sku no puede superar 100 caracteres")
            elif sku in skus_archivo:
                errores_fila.append(f"sku '{sku}' duplicado en el archivo")
            elif sku in skus_bd:
                errores_fila.append(f"sku '{sku}' ya existe en la empresa")

            if not nombre:
                errores_fila.append("nombre es obligatorio")
            elif len(nombre) > 255:
                errores_fila.append("nombre no puede superar 255 caracteres")
            elif nombre in nombres_archivo:
                errores_fila.append(f"nombre '{nombre}' duplicado en el archivo")
            elif nombre in nombres_bd:
                errores_fila.append(f"nombre '{nombre}' ya existe en la empresa")

            unidad_id: int | None = None
            if unidad_raw is None or str(unidad_raw).strip() == "":
                errores_fila.append("unidad_base (unidad_medida_id) es obligatorio")
            else:
                try:
                    unidad_id = int(float(unidad_raw))
                    if unidad_id not in unidades_validas:
                        errores_fila.append(
                            f"unidad_base {unidad_id} no existe o no pertenece a su empresa"
                        )
                except (TypeError, ValueError):
                    errores_fila.append("unidad_base debe ser un número entero (ID de Unidades_medida)")

            tipo_producto_id: int | None = None
            if tipo_raw is not None and str(tipo_raw).strip() != "":
                try:
                    tipo_producto_id = int(float(tipo_raw))
                    if tipo_producto_id not in tipos_validos:
                        errores_fila.append(
                            f"id_tipo_producto {tipo_producto_id} no existe o no pertenece a su empresa"
                        )
                except (TypeError, ValueError):
                    errores_fila.append("id_tipo_producto debe ser un número entero")

            precio_costo: float | None = None
            if precio_raw is not None and str(precio_raw).strip() != "":
                try:
                    precio_costo = float(Decimal(str(precio_raw)))
                    if precio_costo < 0:
                        errores_fila.append("precio_costo no puede ser negativo")
                except (InvalidOperation, ValueError):
                    errores_fila.append("precio_costo debe ser un número válido")

            if errores_fila:
                errores.append({"fila": numero, "sku": sku or None, "errores": errores_fila})
                continue

            skus_archivo.add(sku)
            nombres_archivo.add(nombre)
            validos.append(
                {
                    "fila": numero,
                    "empresa_id": empresa_id,
                    "sku": sku,
                    "nombre": nombre,
                    "unidad_medida_id": unidad_id,
                    "tipo_producto_id": tipo_producto_id,
                    "precio_costo": precio_costo,
                    "activo": True,
                }
            )

        creados = 0
        if validos:
            creados = await self.producto_repo.crear_masivo(
                [
                    {
                        "empresa_id": v["empresa_id"],
                        "sku": v["sku"],
                        "nombre": v["nombre"],
                        "unidad_medida_id": v["unidad_medida_id"],
                        "tipo_producto_id": v["tipo_producto_id"],
                        "precio_costo": v["precio_costo"],
                        "activo": v["activo"],
                    }
                    for v in validos
                ]
            )

        return {
            "total_filas": len(filas),
            "creados": creados,
            "con_errores": len(errores),
            "errores": errores,
        }

    async def _cargar_existentes(self, empresa_id: int) -> tuple[set[str], set[str]]:
        from sqlalchemy import select

        stmt = select(Producto.sku, Producto.nombre).where(Producto.empresa_id == empresa_id)
        result = await self.session.execute(stmt)
        rows = result.all()
        return {r[0] for r in rows if r[0]}, {r[1] for r in rows if r[1]}

    def _parsear_productos(self, contenido: bytes) -> list[dict[str, Any]]:
        wb = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
        if SHEET_PRODUCTOS not in wb.sheetnames:
            raise ValueError(f"Falta la hoja '{SHEET_PRODUCTOS}'")

        ws = wb[SHEET_PRODUCTOS]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        indices = self._indices_columnas(rows[0])
        faltantes = COLUMNAS_OBLIGATORIAS - set(indices.keys())
        if faltantes:
            labels = ", ".join(sorted(faltantes))
            raise ValueError(
                f"Faltan columnas obligatorias en Productos: {labels}. "
                "Use la plantilla actual (unidad_base, sku, nombre)."
            )

        filas: list[dict[str, Any]] = []
        for row_num, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            def val(key: str):
                idx = indices.get(key)
                if idx is None:
                    return None
                return row[idx] if idx < len(row) else None

            sku = self._celda_texto(val("sku"))
            nombre = self._celda_texto(val("nombre"))
            if not sku and not nombre:
                continue

            filas.append(
                {
                    "fila": row_num,
                    "sku": sku,
                    "nombre": nombre,
                    "unidad_medida_id": val("unidad_medida_id"),
                    "tipo_producto_id": val("tipo_producto_id"),
                    "precio_costo": val("precio_costo"),
                }
            )

        wb.close()
        return filas

    def _indices_columnas(self, header_row) -> dict[str, int]:
        indices: dict[str, int] = {}
        for i, cell in enumerate(header_row):
            norm = self._normalizar_header(cell)
            columna = HEADER_A_COLUMNAS.get(norm)
            if columna and columna not in indices:
                indices[columna] = i
        return indices

    @staticmethod
    def _normalizar_header(value) -> str:
        if value is None:
            return ""
        return str(value).strip().lower().replace("_", " ").replace("  ", " ").strip()

    @staticmethod
    def _celda_texto(value) -> str:
        if value is None:
            return ""
        return str(value).strip()
