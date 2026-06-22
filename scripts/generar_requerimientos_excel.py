#!/usr/bin/env python3
"""Genera matriz de requerimientos Khepri Software en Excel."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "docs" / "requerimientos_khepri_software.xlsx"

HEADERS = [
    "R-N°",
    "Nombre del Requerimiento",
    "Tipo Requerimiento\n[Funcional, No Funcional]",
    "Clasificación",
    "Actores Usuarios Relacionados",
    "Descripción corta del requerimiento",
    "Estado",
]

ROWS = [
    (
        "R.01",
        "Autenticación de usuarios",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Operador, Super admin",
        "Permitir iniciar sesión con email y contraseña; emitir JWT con empresa, roles y permisos.",
        "Requerido",
    ),
    (
        "R.02",
        "Recuperación de contraseña",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Operador",
        "Solicitar enlace de restablecimiento por correo (o modo desarrollo en consola); caducidad configurable.",
        "Requerido",
    ),
    (
        "R.03",
        "Bloqueo por intentos fallidos",
        "No Funcional",
        "Seguridad",
        "Administrador, Operador",
        "Tras N intentos fallidos, bloquear cuenta temporalmente; segundo bloqueo permanente.",
        "Requerido",
    ),
    (
        "R.04",
        "Mantenedor de empresas (tenants)",
        "Funcional",
        "Funcional de Administración",
        "Super admin",
        "Registrar, modificar, buscar e inhabilitar empresas multi-tenant; reactivación sin borrar datos.",
        "Requerido",
    ),
    (
        "R.05",
        "Selector de empresa (empresa maestra)",
        "Funcional",
        "Funcional de Usuario",
        "Super admin",
        "Filtrar listados por empresa activa; incluir inhabilitadas solo al seleccionarlas explícitamente.",
        "Requerido",
    ),
    (
        "R.06",
        "Provisionamiento RBAC al crear empresa",
        "Funcional",
        "Funcional de Sistema",
        "Super admin",
        "Al crear empresa, copiar permisos, roles, cargos y vínculos desde plantilla maestra.",
        "Requerido",
    ),
    (
        "R.07",
        "Mantenedor de usuarios",
        "Funcional",
        "Funcional de Administración",
        "Administrador, Super admin",
        "Registrar, modificar, buscar y desactivar usuarios; herencia de roles desde cargo.",
        "Requerido",
    ),
    (
        "R.08",
        "Mantenedor de cargos",
        "Funcional",
        "Funcional de Administración",
        "Administrador, Super admin",
        "Permitir registrar, modificar, buscar y desactivar cargos por empresa.",
        "Requerido",
    ),
    (
        "R.09",
        "Mantenedor de roles y permisos",
        "Funcional",
        "Funcional de Administración",
        "Administrador, Super admin",
        "Gestionar roles, permisos atómicos y asignación rol-permiso por empresa.",
        "Requerido",
    ),
    (
        "R.10",
        "Asignación permisos por cargo",
        "Funcional",
        "Funcional de Administración",
        "Administrador, Super admin",
        "Vincular roles a cargos y sincronizar permisos efectivos de usuarios.",
        "Requerido",
    ),
    (
        "R.11",
        "Mantenedor de bodegas",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Operador, Super admin",
        "Registrar, modificar, buscar y desactivar bodegas asociadas a cada empresa.",
        "Requerido",
    ),
    (
        "R.12",
        "Mantenedor de tipos de zona",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Super admin",
        "Definir tipos de zona (recepción, picking, etc.) por empresa.",
        "Requerido",
    ),
    (
        "R.13",
        "Mantenedor de zonas de bodega",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Operador, Super admin",
        "Registrar ubicaciones/zonas dentro de cada bodega.",
        "Requerido",
    ),
    (
        "R.14",
        "Mantenedor de unidades de medida",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Super admin",
        "Registrar, modificar y buscar unidades de medida del catálogo por empresa.",
        "Requerido",
    ),
    (
        "R.15",
        "Mantenedor de tipos de producto",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Super admin",
        "Clasificar productos por tipo configurable por empresa.",
        "Requerido",
    ),
    (
        "R.16",
        "Mantenedor de productos",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Operador, Super admin",
        "Registrar, modificar, buscar y desactivar productos (SKU, presentaciones, códigos de barras).",
        "Requerido",
    ),
    (
        "R.17",
        "Carga masiva de productos",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Super admin",
        "Importar productos desde plantilla Excel y descargar plantilla de carga.",
        "Requerido",
    ),
    (
        "R.18",
        "Inventario operativo — stock por zona",
        "Funcional",
        "Funcional de Usuario",
        "Operador, Administrador, Super admin",
        "Consultar existencias por bodega, zona y SKU; filtros por empresa.",
        "Requerido",
    ),
    (
        "R.19",
        "Movimientos de inventario",
        "Funcional",
        "Funcional de Usuario",
        "Operador, Administrador",
        "Registrar recepciones, traslados y despachos con trazabilidad y historial.",
        "Requerido",
    ),
    (
        "R.20",
        "Dashboard de inventario",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Operador, Super admin",
        "Visualizar histograma de movimientos y distribución de stock por bodega/ubicación.",
        "Requerido",
    ),
    (
        "R.21",
        "Exportación de datos",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Operador, Super admin",
        "Exportar listados de stock y movimientos a Excel/PDF desde la interfaz.",
        "Requerido",
    ),
    (
        "R.22",
        "Perfil de usuario",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Operador",
        "Consultar y actualizar datos personales del usuario autenticado.",
        "Requerido",
    ),
    (
        "R.23",
        "Landing y acceso público",
        "Funcional",
        "Funcional de Usuario",
        "Visitante, Administrador",
        "Página informativa de Khepri Software y enlace a inicio de sesión.",
        "Requerido",
    ),
    (
        "R.24",
        "Aislamiento multi-tenant",
        "No Funcional",
        "Seguridad",
        "Administrador, Operador, Super admin",
        "Cada empresa accede solo a sus datos; empresa maestra administra tenants autorizados.",
        "Requerido",
    ),
    (
        "R.25",
        "Control de acceso RBAC",
        "No Funcional",
        "Seguridad",
        "Administrador, Super admin",
        "Restringir pantallas y operaciones según permisos del JWT (roles y permisos granulares).",
        "Requerido",
    ),
    (
        "R.26",
        "Correo transaccional",
        "No Funcional",
        "Integración",
        "Administrador, Operador",
        "Envío de correos vía Resend; modo sandbox/desarrollo sin dominio propio.",
        "Requerido",
    ),
    (
        "R.27",
        "Despliegue en nube",
        "No Funcional",
        "Infraestructura",
        "Super admin",
        "Soporte de despliegue en Railway con MySQL, variables de entorno y migraciones SQL.",
        "Requerido",
    ),
    (
        "R.28",
        "Rate limit recuperación contraseña",
        "No Funcional",
        "Seguridad",
        "Visitante",
        "Limitar solicitudes de recuperación por IP y cooldown por usuario para evitar abuso.",
        "Requerido",
    ),
    (
        "R.29",
        "Recepción de mercadería",
        "Funcional",
        "Funcional de Usuario",
        "Operador, Administrador",
        "Registrar ingreso de stock a zona de recepción indicando producto, cantidad y presentación.",
        "Requerido",
    ),
    (
        "R.30",
        "Traslado entre zonas",
        "Funcional",
        "Funcional de Usuario",
        "Operador, Administrador",
        "Mover stock entre zonas de la misma bodega con trazabilidad del movimiento.",
        "Requerido",
    ),
    (
        "R.31",
        "Despacho de mercadería",
        "Funcional",
        "Funcional de Usuario",
        "Operador, Administrador",
        "Registrar salida de stock desde zona de despacho descontando existencias.",
        "Requerido",
    ),
    (
        "R.32",
        "Historial de movimientos auditado",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Operador, Super admin",
        "Consultar historial de recepciones, traslados y despachos con usuario, fecha y cantidades.",
        "Requerido",
    ),
    (
        "R.33",
        "Escaneo con pistola / código SKU",
        "Funcional",
        "Funcional de Usuario",
        "Operador",
        "Seleccionar productos en operaciones de piso mediante escaneo o búsqueda por SKU.",
        "Requerido",
    ),
    (
        "R.34",
        "Presentaciones de producto",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Super admin",
        "Definir presentaciones (caja, pallet, etc.) con factor de conversión a unidad base.",
        "Requerido",
    ),
    (
        "R.35",
        "Configuración zona de recepción por bodega",
        "Funcional",
        "Funcional de Administración",
        "Administrador, Super admin",
        "Asignar zona por defecto para recepciones en cada bodega.",
        "Requerido",
    ),
    (
        "R.36",
        "Cambio de contraseña autenticado",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Operador",
        "Permitir al usuario autenticado cambiar su contraseña validando la actual.",
        "Requerido",
    ),
    (
        "R.37",
        "Asignación de roles por usuario",
        "Funcional",
        "Funcional de Administración",
        "Administrador, Super admin",
        "Consultar y sincronizar roles directos del usuario además de los heredados del cargo.",
        "Requerido",
    ),
    (
        "R.38",
        "Matriz de permisos por rol",
        "Funcional",
        "Funcional de Administración",
        "Administrador, Super admin",
        "Visualizar y editar permisos atómicos asignados a cada rol en interfaz gráfica.",
        "Requerido",
    ),
    (
        "R.39",
        "Paginación, búsqueda y orden en tablas",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Operador, Super admin",
        "Listar registros con paginación server-side, búsqueda textual y orden por columnas.",
        "Requerido",
    ),
    (
        "R.40",
        "Paneles laterales de creación/edición",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Operador, Super admin",
        "Crear y editar entidades en side panels sin abandonar la vista de listado.",
        "Requerido",
    ),
    (
        "R.41",
        "Tema claro y oscuro",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Operador",
        "Alternar apariencia clara/oscuro persistida en preferencias del navegador.",
        "Requerido",
    ),
    (
        "R.42",
        "Notificaciones de operación",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Operador, Super admin",
        "Mostrar mensajes de éxito y error tras acciones CRUD y operaciones de inventario.",
        "Requerido",
    ),
    (
        "R.43",
        "Página de error 404",
        "Funcional",
        "Funcional de Usuario",
        "Visitante, Administrador, Operador",
        "Mostrar página amigable cuando la ruta no existe con enlaces al inicio y login.",
        "Requerido",
    ),
    (
        "R.44",
        "Dashboard general del sistema",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Super admin",
        "Resumen de accesos rápidos, estado del API y contadores básicos por módulo.",
        "Requerido",
    ),
    (
        "R.45",
        "Endpoint provisionar RBAC manual",
        "Funcional",
        "Funcional de Sistema",
        "Super admin",
        "Re-ejecutar copia de permisos/roles/cargos para empresas creadas sin catálogo completo.",
        "Requerido",
    ),
    (
        "R.46",
        "Inhabilitación de empresa sin borrado",
        "Funcional",
        "Funcional de Administración",
        "Super admin",
        "Marcar empresa como inhabilitada ocultando sus datos de listados agregados; reactivable.",
        "Requerido",
    ),
    (
        "R.47",
        "Exportación stock y movimientos",
        "Funcional",
        "Funcional de Usuario",
        "Administrador, Operador, Super admin",
        "Descargar reportes de stock por ubicación e historial en Excel o PDF hasta 50.000 filas.",
        "Requerido",
    ),
    (
        "R.48",
        "Documentación API OpenAPI",
        "No Funcional",
        "Mantenibilidad",
        "Desarrollador, Super admin",
        "Exponer contrato REST en Swagger/ReDoc para integración y pruebas.",
        "Requerido",
    ),
    (
        "R.49",
        "Hash de contraseñas BCrypt",
        "No Funcional",
        "Seguridad",
        "Administrador, Operador",
        "Almacenar contraseñas con hash BCrypt; nunca en texto plano.",
        "Requerido",
    ),
    (
        "R.50",
        "Expiración de tokens JWT",
        "No Funcional",
        "Seguridad",
        "Administrador, Operador",
        "Tokens de acceso con vigencia limitada configurable; requiere nuevo login al expirar.",
        "Requerido",
    ),
]


def main() -> None:
    if len(ROWS) != 50:
        raise SystemExit(f"Se esperaban 50 requerimientos, hay {len(ROWS)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Requerimientos"

    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row_data in enumerate(ROWS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [8, 36, 22, 24, 38, 72, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUTPUT)
        print(f"Generado: {OUTPUT} ({len(ROWS)} requerimientos)")
    except PermissionError:
        alt = OUTPUT.with_name("requerimientos_khepri_software_50.xlsx")
        wb.save(alt)
        print(f"Archivo original en uso. Generado: {alt} ({len(ROWS)} requerimientos)")


if __name__ == "__main__":
    main()
